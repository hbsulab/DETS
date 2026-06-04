from __future__ import annotations

import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "assets" / "data" / "CorrectedData_AfterTimeCorrection.xlsx"
OUT_DIR = ROOT / "assets" / "data" / "aa-exports"

VARIANT_CLASSES = {
    "alpha": "B.1.1.7(Alpha)",
    "beta": "B.1.351(Beta)",
    "delta": "B.1.617.2(Delta)",
    "ba2": "BA.2",
    "jn1": "JN.1",
    "kp2": "KP.2",
}


def normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [normalize(cell) for cell in next(rows_iter)]
    if "class" not in headers:
        raise RuntimeError("Workbook does not contain a 'class' column")
    class_index = headers.index("class")

    rows = list(rows_iter)
    print(f"Loaded {len(rows)} data rows from {WORKBOOK.name}")

    for key, class_label in VARIANT_CLASSES.items():
        out_path = OUT_DIR / f"{key}.csv"
        matched_rows = [row for row in rows if normalize(row[class_index]) == class_label]

        if not matched_rows:
            if out_path.exists():
                out_path.unlink()
            print(f"{key}: no rows for {class_label}")
            continue

        with out_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(matched_rows)

        print(f"{key}: wrote {len(matched_rows)} rows to {out_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
